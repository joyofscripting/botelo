import pathlib
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Union, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import zipfile
from pathlib import Path
from botelo.structures import BoteloProcessBadge
import logging
logger = logging.getLogger(__name__)


class FileFinisherError(Exception):
    """Raised when a file finisher fails.

    Args:
        message (str): Human readable string describing the exception.
        file_finisher (:obj:`AbstractFileFinisher`): An instance of AbstractFileFinisher who raised the exception.

    Attributes:
        message (str): Human readable string describing the exception.
        file_finisher (:obj:`AbstractFileFinisher`): An instance of AbstractFileFinisher who raised the exception."""
    def __init__(self, message, file_finisher):
        self.message = message
        self.file_finisher = file_finisher


class FileFinisherResult(object):
    """Contains the result of the file finisher.

    Args:
        filepath (:obj:`Path`): The path of the finished file.

    Attributes:
        filepath (:obj:`Path`): The path of the finished file.
    """
    def __init__(self, filepath):
        self.filepath = filepath


class AbstractFileFinisher(ABC):
    """This abstract class describes the interface that must be implemented by a file finisher.

    Every data exporter must implement a method named finish_file which returns a FileFinisherResult
    or a FileFinisherError in case of an exception.

    In addition every file finisher must have a property named name which returns its name.
    """
    @abstractmethod
    def finish_file(self, filepath: pathlib.Path, shared_info_collection: dict, process_badge: BoteloProcessBadge) -> Union[FileFinisherResult, FileFinisherError]:
        raise NotImplementedError

    @property
    @abstractmethod
    def name(self) -> str:
        raise NotImplementedError


class TestFileFinisher(AbstractFileFinisher):
    """A file finisher for testing purposes.

    Args:
        name (Optional[str]): The name of this file finisher. Defaults to "Test File Finisher".
        overwrite_file (bool): May an existing file  be overwritten? Defaults to False.
    """
    def __init__(self, name: Optional[str] = "Test File Finisher", overwrite_file: bool = False):
        super().__init__()
        self._name = name
        self.extension = '.test'
        self.overwrite_file = overwrite_file

    @property
    def name(self) -> str:
        """Returns the name of the file finisher."""
        return self._name

    def finish_file(self, filepath: Path, shared_info_collection: dict, process_badge: BoteloProcessBadge) -> Union[FileFinisherResult, FileFinisherError]:
        """Creates a corresponding file in the export folder with the suffix '.test'."""
        try:
            logger.info("Creating a file finisher test file...")
            test_filepath = filepath.with_suffix(self.extension)
            if not self.overwrite_file and test_filepath.exists():
                raise FileExistsError("The filepath already exists: {0}".format(test_filepath))
            with open(test_filepath, 'w') as test_file:
                test_file.write(self._name)
            logger.info("Successfully created a file finisher test file")
            return FileFinisherResult(test_filepath)
        except Exception as e:
            logger.error('{1} Test file finisher failed: {0}'.format(str(e), process_badge.identifier))
            return FileFinisherError(str(e), self)


class ZipFileFinisher(AbstractFileFinisher):
    """A file finisher that creates a ZIP archive of the created export file.

    Args:
        name (Optional[str]): The name of this file finisher. Defaults to "ZIP Archiver".
        overwrite_file (bool): May an existing file  be overwritten? Defaults to False.
    """
    def __init__(self, name: Optional[str] = "ZIP Archiver", overwrite_file: bool = False):
        super().__init__()
        self._name = name
        self.extension = '.zip'
        self.overwrite_file = overwrite_file

    @property
    def name(self):
        """Returns the name of the file finisher."""
        return self._name

    def finish_file(self, filepath: Path, shared_info_collection: dict, process_badge: BoteloProcessBadge) -> Union[FileFinisherResult, FileFinisherError]:
        """Creates a ZIP archive of the given file."""
        try:
            logger.info("{1} Creating ZIP archive for {0}".format(filepath, process_badge.identifier))
            zip_filepath = filepath.with_suffix(self.extension)
            if not self.overwrite_file and zip_filepath.exists():
                raise FileExistsError("The ZIP filepath already exists: {0}".format(zip_filepath))
            with zipfile.ZipFile(zip_filepath, mode='w') as zf:
                zf.write(filepath, compress_type=zipfile.ZIP_DEFLATED, arcname=filepath.name)
            logger.info("{0} Successfully created ZIP archive".format(process_badge))
            return FileFinisherResult(zip_filepath)
        except Exception as e:
            logger.error('{1} Zip archive creation failed: {0}'.format(str(e), process_badge.identifier))
            return FileFinisherError(str(e), self)


class ExcelDecorationElement(object):
    """Stores the label and content for a single Excel decoration element.

    Args:
        label (str): The label of the decoration element, e.g. 'Created by'.
        content (str): The content of the decoration element, e.g. 'Darth Vader'.
    """
    def __init__(self, label: str, content: str):
        self.label = label
        self.content = content


class ExcelDecoration(object):
    """Stores all elements for an Excel decoration.

    Args:
        sheet_name (str): The name of the worksheet to be created.
        title (str): The title of the decoration, e.g. 'Details' or 'Info'.
    """
    def __init__(self, sheet_name: str, title: str):
        self.sheet_name = sheet_name
        self.title = title
        self.elements = []

    def add_element(self, element):
        """Adds an excel decoration element."""
        self.elements.append(element)


class ExcelDecorator(AbstractFileFinisher):
    """Finishes an Excel file with one or more decorations.

    Args:
        name (Optional[str]): The name of this file finisher. Defaults to "Excel Decorator".
    """
    def __init__(self, name: Optional[str] = "Excel Decorator"):
        super().__init__()
        self._name = name

    @property
    def name(self) -> str:
        """Returns the name of the file finisher."""
        return self._name

    def _replace_placeholders(self, content):
        if content == 'CURRENT_DATETIME':
            return datetime.now()
        else:
            return content

    def finish_file(self, filepath: Path, shared_info_collection: dict, process_badge: BoteloProcessBadge) -> Union[FileFinisherResult, FileFinisherError]:
        """Finishes an Excel file with decorations."""
        try:
            title_font = Font(name='Calibri',
                              size=22,
                              bold=False,
                              italic=False,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            key_font = Font(name='Calibri',
                            size=11,
                            bold=True,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000')

            value_font = Font(name='Calibri',
                              size=11,
                              bold=False,
                              italic=False,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            logger.info("{1} Writing additional informations to Excel file {0}...".format(filepath, process_badge.identifier))
            wb = load_workbook(filename=filepath)

            for decoration in self.get_decorations(shared_info_collection):
                logger.info("{2} Adding work sheet named '{0}' with title '{1}'".format(decoration.sheet_name, decoration.title, process_badge.identifier))
                ws = wb.create_sheet(title=decoration.sheet_name)
                ws['B3'] = decoration.title
                ws['B3'].font = title_font

                cell_index = 6
                for element in decoration.elements:
                    key_cell = "B{0}".format(cell_index)
                    value_cell = "C{0}".format(cell_index)

                    ws[key_cell] = element.label
                    ws[key_cell].font = key_font
                    ws[key_cell].alignment = Alignment(horizontal='right')
                    ws[value_cell] = self._replace_placeholders(element.content)
                    ws[value_cell].font = value_font
                    ws[value_cell].alignment = Alignment(horizontal='left')

                    cell_index += 1

            logger.info("{1} Saving Excel file {0}...".format(filepath, process_badge.identifier))
            wb.save(filepath)

            logger.info("{0} Successfully decorated the Excel file".format(process_badge.identifier))

            return FileFinisherResult(filepath)
        except Exception as e:
            logger.error('{1} Excel decoration failed: {0}'.format(str(e), process_badge.identifier))
            return FileFinisherError(str(e), self)

    def get_decorations(self, shared_info_collection):
        excel_decorations = []

        excel_info_decoration = ExcelDecoration('Info', shared_info_collection['jira_issue_title'])
        excel_info_decoration.add_element(ExcelDecorationElement('Erstellt am', 'CURRENT_DATETIME'))
        excel_info_decoration.add_element(ExcelDecorationElement('Erstellt von', shared_info_collection['creator']))
        excel_info_decoration.add_element(ExcelDecorationElement('', ''))
        excel_info_decoration.add_element(ExcelDecorationElement('Server', shared_info_collection['db_host']))
        excel_info_decoration.add_element(ExcelDecorationElement('Datenbank', shared_info_collection['db_name']))
        excel_info_decoration.add_element(ExcelDecorationElement('Jira-URL', shared_info_collection['jira_url']))

        excel_decorations.append(excel_info_decoration)

        excel_info_decoration = ExcelDecoration('SQL', 'Abfragedetails')
        excel_info_decoration.add_element(ExcelDecorationElement('Abfragedauer', shared_info_collection['BT_DATA_COLLECTION_DURATION']))
        excel_info_decoration.add_element(ExcelDecorationElement('SQL-Abfrage', shared_info_collection['BT_QUERY']))

        excel_decorations.append(excel_info_decoration)

        return excel_decorations
